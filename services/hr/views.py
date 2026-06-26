import csv
import io
import openpyxl
from django.utils import timezone
from django.http import HttpResponse
from datetime import datetime
from django.db.models import Q
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from shared.permissions import make_module_permission, user_can, P

HRPermission = make_module_permission(P.HR_VIEW, P.HR_MANAGE)
# Claims are self-service: any employee with HR access manages their own claims
# (querysets scoped to the owner); approve/reject is gated to the supervisor.
ClaimPermission = make_module_permission(P.HR_VIEW)
from .models import Employee, LeaveType, LeaveBalance, LeaveApplication, Attendance, Certification, PublicHoliday, WorkSchedule, ManpowerSettings, StaffDeployment, PersonalGoal, Claim, ClaimItem, ClaimAttachment
from .serializers import (
    EmployeeSerializer, EmployeeTreeSerializer, LeaveTypeSerializer, LeaveBalanceSerializer,
    LeaveApplicationSerializer, AttendanceSerializer, CertificationSerializer, PublicHolidaySerializer,
    ClockInResponseSerializer, WorkScheduleSerializer, ManpowerSettingsSerializer, StaffDeploymentSerializer,
    PersonalGoalSerializer, ClaimSerializer, ClaimItemSerializer, ClaimAttachmentSerializer,
)
from .permissions import IsClockInAllowed
from shared.utils import haversine_distance
import os as _os

def fb_read_csv(path):
    if not _os.path.exists(path):
        return []
    with open(path, newline='', encoding='utf-8') as f:
        return list(csv.DictReader(f))

def fb_write_csv(path, rows, fields):
    _os.makedirs(_os.path.dirname(path), exist_ok=True)
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
        w.writeheader()
        w.writerows(rows)


class TenantScopedMixin:
    permission_classes = [HRPermission]

    def get_queryset(self):
        return self.queryset.filter(is_active=True)

    def perform_create(self, serializer):
        serializer.save()


class EmployeeViewSet(TenantScopedMixin, viewsets.ModelViewSet):
    queryset = Employee.objects.select_related('department', 'position')
    serializer_class = EmployeeSerializer
    pagination_class = None

    def get_queryset(self):
        qs = super().get_queryset()
        can_clock_in = self.request.query_params.get('can_clock_in')
        if can_clock_in == 'true':
            qs = qs.filter(can_clock_in=True)
        return qs

    @action(detail=False, methods=['get'], url_path='me')
    def me(self, request):
        try:
            emp = Employee.objects.select_related('department', 'position').get(user=request.user)
        except Employee.DoesNotExist:
            return Response({'employee': None, 'leave_balances': [], 'leave_applications': []})

        balances = LeaveBalance.objects.filter(employee=emp).select_related('leave_type')
        applications = LeaveApplication.objects.filter(employee=emp).select_related('leave_type').order_by('-created_at')[:5]

        today = timezone.now().date()
        attendance = Attendance.objects.filter(employee=emp, date=today).first()

        return Response({
            'employee': EmployeeSerializer(emp, context={'request': request}).data,
            'leave_balances': LeaveBalanceSerializer(balances, many=True).data,
            'leave_applications': LeaveApplicationSerializer(applications, many=True).data,
            'today_attendance': AttendanceSerializer(attendance).data if attendance else None,
        })


class LeaveTypeViewSet(TenantScopedMixin, viewsets.ModelViewSet):
    queryset = LeaveType.objects.all()
    serializer_class = LeaveTypeSerializer


class LeaveBalanceViewSet(TenantScopedMixin, viewsets.ModelViewSet):
    queryset = LeaveBalance.objects.select_related('employee', 'leave_type')
    serializer_class = LeaveBalanceSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        employee_id = self.request.query_params.get('employee')
        if employee_id:
            qs = qs.filter(employee_id=employee_id)
        return qs


class LeaveApplicationViewSet(TenantScopedMixin, viewsets.ModelViewSet):
    queryset = LeaveApplication.objects.select_related('employee', 'leave_type', 'approved_by')
    serializer_class = LeaveApplicationSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        employee_id = self.request.query_params.get('employee')
        status = self.request.query_params.get('status')
        if employee_id:
            qs = qs.filter(employee_id=employee_id)
        if status:
            qs = qs.filter(status=status)
        return qs.order_by('-created_at')

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        if not user_can(request.user, P.HR_APPROVE_LEAVE):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        leave = self.get_object()
        leave.status = 'approved'
        leave.approved_by = request.user
        leave.approved_at = timezone.now()
        leave.remarks = request.data.get('remarks', '')
        leave.save()
        return Response(LeaveApplicationSerializer(leave).data)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        if not user_can(request.user, P.HR_APPROVE_LEAVE):
            return Response({'detail': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        leave = self.get_object()
        leave.status = 'rejected'
        leave.approved_by = request.user
        leave.approved_at = timezone.now()
        leave.remarks = request.data.get('remarks', '')
        leave.save()
        return Response(LeaveApplicationSerializer(leave).data)


class AttendanceViewSet(TenantScopedMixin, viewsets.ModelViewSet):
    queryset = Attendance.objects.select_related('employee')
    serializer_class = AttendanceSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        employee_id = self.request.query_params.get('employee')
        if employee_id:
            qs = qs.filter(employee_id=employee_id)
        return qs.order_by('-date')

    @action(detail=False, methods=['post'], permission_classes=[IsClockInAllowed])
    def clock_in(self, request):
        try:
            employee = request.user.employee_profile
            today = timezone.now().date()
            today_str = today.strftime('%d-%m-%Y')

            # Check schedule exists in CSV
            rows = fb_read_csv(CSV_PATH)
            schedule = next((r for r in rows if r.get('emp_no') == employee.emp_no and r.get('date') == today_str), None)
            if not schedule:
                return Response(
                    {'success': False, 'message': 'No schedule assigned for today. Contact your supervisor.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Check GPS within allowed radius
            gps_lat = request.data.get('gps_lat')
            gps_lng = request.data.get('gps_lng')
            project_id = request.data.get('project_id')
            outside_geofence = False
            if gps_lat and gps_lng:
                distance = haversine_distance(gps_lat, gps_lng, schedule['location_lat'], schedule['location_lng'])
                radius = int(schedule.get('radius', 200))
                if distance > radius:
                    if not project_id:
                        return Response({
                            'success': False,
                            'message': f'You are {int(distance)}m away from {schedule["location_name"]}. Select a project to clock in remotely.'
                        }, status=status.HTTP_400_BAD_REQUEST)
                    outside_geofence = True

            # Flag late if past shift_start
            now = timezone.now()
            from datetime import time as time_type
            shift_start = datetime.strptime(schedule.get('shift_start', '00:00'), '%H:%M').time()
            attendance_status = 'late' if timezone.localtime(now).time() > shift_start else 'present'

            record, created = Attendance.objects.get_or_create(
                employee=employee,
                date=today,
            )

            record.clock_in = now
            record.clock_in_photo = request.FILES.get('photo')
            record.status = attendance_status

            if gps_lat and gps_lng:
                record.clock_in_gps = {'lat': float(gps_lat), 'lng': float(gps_lng)}

            address = request.data.get('address')
            if address:
                record.clock_in_address = address

            if project_id and outside_geofence:
                from services.projects.models import Project
                try:
                    record.project = Project.objects.get(id=project_id)
                except Project.DoesNotExist:
                    pass

            record.save()

            photo_url = record.clock_in_photo.url if record.clock_in_photo else None
            late_note = ' (Late)' if attendance_status == 'late' else ''
            response = {
                'success': True,
                'message': f'Clocked in at {timezone.localtime(now).strftime("%H:%M:%S")}{late_note}',
                'clock_in_time': now,
                'status': attendance_status,
                'photo_url': photo_url,
                'gps_location': record.clock_in_gps,
                'schedule': {
                    'location': schedule['location_name'],
                    'shift_start': schedule['shift_start'],
                    'shift_end': schedule['shift_end'],
                },
            }
            return Response(response, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {'success': False, 'message': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['post'], permission_classes=[IsClockInAllowed])
    def clock_out(self, request):
        try:
            employee = request.user.employee_profile
            today = timezone.now().date()

            record = Attendance.objects.get(employee=employee, date=today)

            now = timezone.now()
            record.clock_out = now
            record.clock_out_photo = request.FILES.get('photo')

            gps_lat = request.data.get('gps_lat')
            gps_lng = request.data.get('gps_lng')
            if gps_lat and gps_lng:
                record.clock_out_gps = {'lat': float(gps_lat), 'lng': float(gps_lng)}

            address = request.data.get('address')
            if address:
                record.clock_out_address = address

            if record.clock_in:
                delta = now - record.clock_in
                hours = delta.total_seconds() / 3600
                record.hours = round(hours, 2)

            record.save()

            photo_url = record.clock_out_photo.url if record.clock_out_photo else None
            response = {
                'success': True,
                'message': f'Clocked out at {timezone.localtime(now).strftime("%H:%M:%S")}',
                'clock_out_time': now,
                'hours_worked': record.hours,
                'photo_url': photo_url,
                'gps_location': record.clock_out_gps,
            }
            return Response(response, status=status.HTTP_200_OK)
        except Attendance.DoesNotExist:
            return Response(
                {'success': False, 'message': 'No clock-in record for today'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'success': False, 'message': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


CSV_PATH = '/mnt/data/1os/database/Work_schedule.csv'


def _try_parse(val):
    try:
        return _parse_date(val)
    except Exception:
        return None
CSV_FIELDS = ['emp_no', 'first_name', 'last_name', 'date', 'shift_start', 'shift_end', 'location_name', 'location_lat', 'location_lng', 'radius']


def _parse_date(val):
    """Parse date string accepting DD-MM-YYYY, DD/MM/YYYY, YYYY-MM-DD."""
    s = str(val).split(' ')[0].split('T')[0]
    for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f'Invalid date: {val}')


def _row_clock_status(emp, date_obj):
    record = Attendance.objects.filter(employee=emp, date=date_obj).first()
    if not record or not record.clock_in:
        return ('Missed' if date_obj < timezone.now().date() else 'Pending'), None
    return ('Late' if record.status == 'late' else 'Done'), timezone.localtime(record.clock_in).strftime('%H:%M')


def _enrich(row):
    """Add id, employee_name, clock_status, clock_in_time to a CSV row dict."""
    try:
        date_obj = _parse_date(row.get('date', ''))
    except ValueError:
        row['id'] = f"{row.get('emp_no','')}_{row.get('date','')}"
        row['employee_name'] = f"{row.get('first_name','')} {row.get('last_name','')}".strip()
        row['clock_status'] = 'Pending'
        row['clock_in_time'] = None
        return row

    emp = Employee.objects.filter(emp_no=row.get('emp_no', '')).first()
    row['id'] = f"{row['emp_no']}_{date_obj.strftime('%d-%m-%Y')}"
    row['employee_id'] = str(emp.id) if emp else None
    row['employee_name'] = emp.full_name if emp else f"{row.get('first_name','')} {row.get('last_name','')}".strip()
    clock_status, clock_time = _row_clock_status(emp, date_obj) if emp else ('Pending', None)
    row['clock_status'] = clock_status
    row['clock_in_time'] = clock_time
    return row


class WorkScheduleViewSet(viewsets.ViewSet):
    permission_classes = [HRPermission]

    def _all_emp_nos(self):
        return set(Employee.objects.filter(is_active=True).values_list('emp_no', flat=True))

    def list(self, request):
        emp_nos = self._all_emp_nos()
        rows = [r for r in fb_read_csv(CSV_PATH) if r.get('emp_no', '') in emp_nos]

        date_filter = request.query_params.get('date')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        emp_filter = request.query_params.get('employee')

        month_filter = request.query_params.get('month')  # YYYY-MM
        if month_filter:
            try:
                y, m = month_filter.split('-')
                rows = [r for r in rows if r.get('date', '').endswith(f'-{m}-{y}') or
                        (lambda d: d.year == int(y) and d.month == int(m) if d else False)(_try_parse(r.get('date', '')))]
            except Exception:
                pass

        # Date range filtering
        if date_from or date_to:
            try:
                from_date = _parse_date(date_from) if date_from else None
                to_date = _parse_date(date_to) if date_to else None
                filtered = []
                for r in rows:
                    try:
                        r_date = _parse_date(r.get('date', ''))
                        if from_date and r_date < from_date:
                            continue
                        if to_date and r_date > to_date:
                            continue
                        filtered.append(r)
                    except ValueError:
                        pass
                rows = filtered
            except Exception:
                pass

        if date_filter:
            rows = [r for r in rows if r.get('date') == date_filter or
                    (lambda d: d.strftime('%d-%m-%Y') == date_filter if d else False)(_try_parse(r.get('date', '')))]
        if emp_filter:
            emp = Employee.objects.filter(id=emp_filter).first()
            if emp:
                rows = [r for r in rows if r.get('emp_no') == emp.emp_no]

        rows = sorted(rows, key=lambda r: (r.get('date', ''), r.get('shift_start', '')))
        rows = [_enrich(r) for r in rows]
        return Response({'results': rows, 'count': len(rows)})

    def create(self, request):
        emp_id = request.data.get('employee')
        emp = Employee.objects.filter(id=emp_id).first() if emp_id else None
        if not emp:
            return Response({'error': 'Employee not found'}, status=400)

        try:
            date_obj = _parse_date(request.data.get('date', ''))
        except ValueError:
            return Response({'error': 'Invalid date format'}, status=400)

        date_str = date_obj.strftime('%d-%m-%Y')
        rows = fb_read_csv(CSV_PATH)

        if any(r.get('emp_no') == emp.emp_no and r.get('date') == date_str for r in rows):
            return Response({'error': f'Schedule already exists for {emp.emp_no} on {date_str}'}, status=400)

        new_row = {
            'emp_no': emp.emp_no,
            'first_name': emp.first_name,
            'last_name': emp.last_name,
            'date': date_str,
            'shift_start': request.data.get('shift_start', ''),
            'shift_end': request.data.get('shift_end', ''),
            'location_name': request.data.get('location_name', ''),
            'location_lat': request.data.get('location_lat', ''),
            'location_lng': request.data.get('location_lng', ''),
            'radius': request.data.get('radius', 200),
        }
        rows.append(new_row)
        fb_write_csv(CSV_PATH, rows, CSV_FIELDS)
        return Response(_enrich(new_row), status=status.HTTP_201_CREATED)

    def update(self, request, pk=None):
        emp_no, date_str = pk.rsplit('_', 1)
        rows = fb_read_csv(CSV_PATH)
        updated = None
        for i, r in enumerate(rows):
            if r.get('emp_no') == emp_no and r.get('date') == date_str:
                rows[i].update({
                    'shift_start': request.data.get('shift_start', r['shift_start']),
                    'shift_end': request.data.get('shift_end', r['shift_end']),
                    'location_name': request.data.get('location_name', r['location_name']),
                    'location_lat': request.data.get('location_lat', r['location_lat']),
                    'location_lng': request.data.get('location_lng', r['location_lng']),
                    'radius': request.data.get('radius', r['radius']),
                })
                updated = rows[i]
                break
        if not updated:
            return Response({'error': 'Schedule not found'}, status=404)
        fb_write_csv(CSV_PATH, rows, CSV_FIELDS)
        return Response(_enrich(updated))

    def destroy(self, request, pk=None):
        emp_no, date_str = pk.rsplit('_', 1)
        rows = fb_read_csv(CSV_PATH)
        rows = [r for r in rows if not (r.get('emp_no') == emp_no and r.get('date') == date_str)]
        fb_write_csv(CSV_PATH, rows, CSV_FIELDS)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        emp_nos = self._all_emp_nos()
        rows = [r for r in fb_read_csv(CSV_PATH) if r.get('emp_no', '') in emp_nos]
        rows = [_enrich(r) for r in rows]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Schedules'
        headers = CSV_FIELDS + ['clock_in_status', 'clock_in_time']
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, '') for h in headers])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="Work_schedule.xlsx"'
        return resp

    @action(detail=False, methods=['post'])
    def import_file(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'success': False, 'message': 'No file uploaded'}, status=400)

        incoming = []
        filename = file.name.lower()
        try:
            if filename.endswith('.xlsx'):
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(v for v in row):
                        incoming.append(dict(zip(headers, [str(v) if v is not None else '' for v in row])))
            elif filename.endswith('.csv'):
                text = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(text))
                incoming = [r for r in reader if any(str(v).strip() for v in r.values())]
            else:
                return Response({'success': False, 'message': 'Only .csv or .xlsx files are supported'}, status=400)
        except Exception as e:
            return Response({'success': False, 'message': f'File parse error: {e}'}, status=400)

        if not incoming:
            return Response({'success': False, 'message': 'File is empty'}, status=400)

        emp_nos = self._all_emp_nos()
        existing = fb_read_csv(CSV_PATH)
        existing_keys = {(r.get('emp_no'), r.get('date')) for r in existing}

        errors = []
        validated = []
        for i, row in enumerate(incoming, start=2):
            emp_no = str(row.get('emp_no', '')).strip()
            date_val = str(row.get('date', '')).strip()
            shift_start = str(row.get('shift_start', '')).strip()
            shift_end = str(row.get('shift_end', '')).strip()
            location_name = str(row.get('location_name', '')).strip()
            location_lat = str(row.get('location_lat', '')).strip()
            location_lng = str(row.get('location_lng', '')).strip()
            radius = str(row.get('radius', '200')).strip() or '200'

            if not all([emp_no, date_val, shift_start, shift_end, location_name, location_lat, location_lng]):
                errors.append(f'Row {i}: Missing required fields')
                continue
            if emp_no not in emp_nos:
                errors.append(f'Row {i}: Employee {emp_no} not found')
                continue
            try:
                date_obj = _parse_date(date_val)
                date_str = date_obj.strftime('%d-%m-%Y')
            except ValueError:
                errors.append(f'Row {i}: Invalid date "{date_val}" — use DD-MM-YYYY')
                continue
            if (emp_no, date_str) in existing_keys:
                errors.append(f'Row {i}: Schedule already exists for {emp_no} on {date_str}')
                continue

            emp = Employee.objects.filter(emp_no=emp_no).first()
            validated.append({
                'emp_no': emp_no,
                'first_name': emp.first_name if emp else '',
                'last_name': emp.last_name if emp else '',
                'date': date_str,
                'shift_start': shift_start,
                'shift_end': shift_end,
                'location_name': location_name,
                'location_lat': location_lat,
                'location_lng': location_lng,
                'radius': radius,
            })

        if errors:
            return Response({'success': False, 'errors': errors, 'message': f'{len(errors)} error(s) — nothing imported'}, status=400)

        fb_write_csv(CSV_PATH, existing + validated, CSV_FIELDS)
        return Response({'success': True, 'message': f'{len(validated)} schedule(s) imported'})


class CertificationViewSet(TenantScopedMixin, viewsets.ModelViewSet):
    queryset = Certification.objects.select_related('employee')
    serializer_class = CertificationSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        employee_id = self.request.query_params.get('employee')
        if employee_id:
            qs = qs.filter(employee_id=employee_id)
        return qs.order_by('-issue_date')


class PublicHolidayViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = PublicHoliday.objects.all()
    serializer_class = PublicHolidaySerializer
    permission_classes = [HRPermission]


@api_view(['GET'])
@permission_classes([HRPermission])
def org_tree(request):
    """Return recursive org chart tree from root(s)."""
    roots = Employee.objects.filter(is_active=True, manager__isnull=True)
    data = EmployeeTreeSerializer(roots, many=True, context={'request': request}).data
    if len(data) == 1:
        return Response(data[0])
    return Response({'id': None, 'full_name': 'Company', 'children': data})


@api_view(['GET'])
@permission_classes([HRPermission])
def employee_me(request):
    """Return the employee profile for the logged-in user."""
    try:
        employee = Employee.objects.select_related('department', 'position').get(
            user=request.user
        )
        return Response(EmployeeSerializer(employee).data)
    except Employee.DoesNotExist:
        return Response({'detail': 'No employee profile found.'}, status=404)


class StaffDeploymentViewSet(TenantScopedMixin, viewsets.ModelViewSet):
    queryset = StaffDeployment.objects.select_related('employee')
    serializer_class = StaffDeploymentSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        emp = self.request.query_params.get('employee')
        if emp:
            qs = qs.filter(employee_id=emp)
        return qs.order_by('-date_from')

    @action(detail=True, methods=['post'], url_path='generate')
    def generate_schedules(self, request, pk=None):
        """Expand this deployment into WorkSchedule CSV rows for each matching date."""
        import datetime as dt
        dep = self.get_object()
        emp = dep.employee

        rows = fb_read_csv(CSV_PATH)
        existing_keys = {(r.get('emp_no'), r.get('date')) for r in rows}

        created, skipped = 0, 0
        current = dep.date_from
        new_rows = []
        while current <= dep.date_to:
            if current.weekday() in dep.days_of_week:
                date_str = current.strftime('%d-%m-%Y')
                key = (emp.emp_no, date_str)
                if key not in existing_keys:
                    new_rows.append({
                        'emp_no': emp.emp_no,
                        'first_name': emp.first_name,
                        'last_name': emp.last_name,
                        'date': date_str,
                        'shift_start': dep.shift_start.strftime('%H:%M'),
                        'shift_end': dep.shift_end.strftime('%H:%M'),
                        'location_name': dep.location_name,
                        'location_lat': str(dep.location_lat),
                        'location_lng': str(dep.location_lng),
                        'radius': str(dep.radius),
                    })
                    existing_keys.add(key)
                    created += 1
                else:
                    skipped += 1
            current += dt.timedelta(days=1)

        if new_rows:
            fb_write_csv(CSV_PATH, rows + new_rows, CSV_FIELDS)
        return Response({'created': created, 'skipped': skipped,
                         'message': f'{created} schedule(s) generated, {skipped} already existed.'})


class ManpowerSettingsViewSet(viewsets.ModelViewSet):
    serializer_class = ManpowerSettingsSerializer
    permission_classes = [HRPermission]
    queryset = ManpowerSettings.objects.all()

    def get_queryset(self):
        return ManpowerSettings.objects.all()

    def perform_create(self, serializer):
        serializer.save()

    @action(detail=False, methods=['get', 'post'], url_path='current')
    def current(self, request):
        """Get or create/update manpower settings."""
        obj = ManpowerSettings.objects.first()
        if obj is None:
            obj = ManpowerSettings.objects.create()
        if request.method == 'POST':
            serializer = self.get_serializer(obj, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(serializer.data)
        return Response(self.get_serializer(obj).data)


class PersonalGoalViewSet(viewsets.ModelViewSet):
    permission_classes = [HRPermission]
    serializer_class = PersonalGoalSerializer

    def get_queryset(self):
        return PersonalGoal.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


# ---------------------------------------------------------------------------
# Monthly expense claims (self-service; approved by reporting supervisor)
# ---------------------------------------------------------------------------
class ClaimViewSet(viewsets.ModelViewSet):
    permission_classes = [ClaimPermission]
    serializer_class = ClaimSerializer

    def get_queryset(self):
        user = self.request.user
        qs = (Claim.objects
              .select_related('claimant', 'approver', 'reviewed_by')
              .prefetch_related('items__attachments'))
        # Base visibility: own claims + claims awaiting my approval (HR_MANAGE sees all)
        if not user_can(user, P.HR_MANAGE):
            qs = qs.filter(Q(claimant=user) | Q(approver=user))
        scope = self.request.query_params.get('scope')
        status_f = self.request.query_params.get('status')
        if scope == 'mine':
            qs = qs.filter(claimant=user)
        elif scope == 'to_approve':
            qs = qs.filter(approver=user)
        if status_f:
            qs = qs.filter(status=status_f)
        return qs.distinct()

    def perform_create(self, serializer):
        serializer.save(claimant=self.request.user)

    def _owner_editable(self, claim):
        return claim.claimant_id == self.request.user.id and claim.status in ('draft', 'rejected')

    def update(self, request, *args, **kwargs):
        if not self._owner_editable(self.get_object()):
            raise PermissionDenied('Only your own draft/rejected claims can be edited.')
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not self._owner_editable(self.get_object()):
            raise PermissionDenied('Only your own draft/rejected claims can be deleted.')
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        claim = self.get_object()
        if claim.claimant_id != request.user.id:
            raise PermissionDenied('Not your claim.')
        if claim.status not in ('draft', 'rejected'):
            return Response({'detail': 'Claim has already been submitted.'}, status=status.HTTP_400_BAD_REQUEST)
        if not claim.items.exists():
            return Response({'detail': 'Add at least one item before submitting.'}, status=status.HTTP_400_BAD_REQUEST)
        emp = getattr(request.user, 'employee_profile', None)
        supervisor = emp.manager if emp else None
        approver_user = supervisor.user if (supervisor and supervisor.user_id) else None
        if not approver_user:
            return Response(
                {'detail': 'No reporting supervisor assigned — contact HR.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        claim.approver = approver_user
        claim.status = 'submitted'
        claim.submitted_at = timezone.now()
        claim.reviewed_by = None
        claim.reviewed_at = None
        claim.remarks = ''
        claim.save()
        claim.recalculate_total()
        return Response(ClaimSerializer(claim).data)

    def _can_review(self, claim, user):
        return claim.approver_id == user.id or user_can(user, P.HR_MANAGE)

    def _review(self, request, new_status):
        claim = self.get_object()
        if not self._can_review(claim, request.user):
            raise PermissionDenied('Only the reporting supervisor can review this claim.')
        if claim.status != 'submitted':
            return Response({'detail': 'Only submitted claims can be reviewed.'}, status=status.HTTP_400_BAD_REQUEST)
        claim.status = new_status
        claim.reviewed_by = request.user
        claim.reviewed_at = timezone.now()
        claim.remarks = request.data.get('remarks', '')
        claim.save()
        return Response(ClaimSerializer(claim).data)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        return self._review(request, 'approved')

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        return self._review(request, 'rejected')


class ClaimItemViewSet(viewsets.ModelViewSet):
    permission_classes = [ClaimPermission]
    serializer_class = ClaimItemSerializer

    def get_queryset(self):
        qs = (ClaimItem.objects
              .select_related('claim')
              .prefetch_related('attachments')
              .filter(claim__claimant=self.request.user))
        claim_id = self.request.query_params.get('claim')
        if claim_id:
            qs = qs.filter(claim_id=claim_id)
        return qs

    def _guard(self, claim):
        if claim.claimant_id != self.request.user.id:
            raise PermissionDenied('Not your claim.')
        if claim.status not in ('draft', 'rejected'):
            raise PermissionDenied('Cannot edit items of a submitted claim.')

    def perform_create(self, serializer):
        self._guard(serializer.validated_data['claim'])
        serializer.save()

    def perform_update(self, serializer):
        self._guard(serializer.instance.claim)
        serializer.save()

    def perform_destroy(self, instance):
        self._guard(instance.claim)
        instance.delete()


class ClaimAttachmentViewSet(viewsets.ModelViewSet):
    permission_classes = [ClaimPermission]
    serializer_class = ClaimAttachmentSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        qs = (ClaimAttachment.objects
              .select_related('item__claim')
              .filter(item__claim__claimant=self.request.user))
        item_id = self.request.query_params.get('item')
        if item_id:
            qs = qs.filter(item_id=item_id)
        return qs

    def _guard(self, item):
        if item.claim.claimant_id != self.request.user.id:
            raise PermissionDenied('Not your claim.')
        if item.claim.status not in ('draft', 'rejected'):
            raise PermissionDenied('Cannot change attachments of a submitted claim.')

    def perform_create(self, serializer):
        self._guard(serializer.validated_data['item'])
        serializer.save(uploaded_by=self.request.user)

    def perform_destroy(self, instance):
        self._guard(instance.item)
        instance.delete()